"""
question_loader.py — Parse and validate question CSV uploads.
Works for ANY course type, not just LSAT.

New in this version:
  - content_hash is computed for every row (duplicate content detection)
  - uploaded question_id values are ignored; the database assigns course IDs
  - process_upload() returns a rich dict instead of a 3-tuple
"""

import hashlib
import random
import pandas as pd
import io
from src.database import insert_questions, get_course

# Columns that MUST be in the CSV
REQUIRED_COLS = {
    "stimulus", "question_type",
}

# All recognised columns (optional ones get empty string if missing)
ALL_COLS = [
    "question_id", "section_type", "question_type", "difficulty",
    "passage", "stimulus",
    "choice_a", "choice_b", "choice_c", "choice_d", "choice_e",
    "correct_answer", "explanation",
    "wrong_answer_a", "wrong_answer_b", "wrong_answer_c",
    "wrong_answer_d", "wrong_answer_e",
    "source", "tags",
]

VALID_ANSWERS = {"A", "B", "C", "D", "E"}
OPEN_ENDED_TYPES = {"open-ended", "open ended", "open_ended", "written response", "free response"}
MULTIPLE_CHOICE_TYPES = {"multiple choice", "multiple-choice", "mc", "mcq"}


def _is_alpha_street_source(source: str) -> bool:
    normalized = "".join(ch for ch in str(source or "").lower() if ch.isalnum())
    return "alphastreet" in normalized


def _cfa_level_label(*values: str) -> str:
    roman_levels = {"i": "1", "ii": "2", "iii": "3"}

    for value in values:
        text = str(value or "")
        normalized = " ".join(text.replace("-", " ").split())
        lower = normalized.lower()
        marker = "cfa level "
        if marker not in lower:
            continue

        after_marker = normalized[lower.index(marker) + len(marker):].strip()
        if not after_marker:
            continue

        level_token = after_marker.split()[0].strip(":,;()[]{}")
        level = roman_levels.get(level_token.lower(), level_token)
        if level in {"1", "2", "3"}:
            return f"CFA Level {level}"

    return "CFA"


def _apply_source_module_rules(rows: list[dict], course_id=None) -> None:
    course_title = ""
    if course_id is not None:
        course = get_course(course_id)
        course_title = (course or {}).get("title", "")

    for row in rows:
        if not _is_alpha_street_source(row.get("source", "")):
            continue

        level_label = _cfa_level_label(course_title, row.get("section_type", ""))
        row["section_type"] = f"Alpha Street {level_label}"


def is_open_ended_question(q: dict) -> bool:
    """Return True when a row/question should be rendered as written response."""
    if q.get("_force_open_ended"):
        return True
    qtype = str(q.get("question_type") or "").strip().lower()
    return qtype in OPEN_ENDED_TYPES


def _shuffle_multiple_choice_row(r: dict) -> None:
    """Shuffle A-E choices and update the correct answer letter."""
    correct = str(r.get("correct_answer") or "").strip().upper()
    if correct not in VALID_ANSWERS:
        return

    items = [
        (letter, str(r.get(f"choice_{letter.lower()}", "") or "").strip())
        for letter in ["A", "B", "C", "D", "E"]
    ]
    if any(not text for _, text in items):
        return

    correct_text = dict(items)[correct]
    seed_src = _content_hash(
        r.get("stimulus", ""),
        *(text for _, text in items),
        correct,
    )
    rng = random.Random(seed_src)
    rng.shuffle(items)

    for new_letter, (_, text) in zip(["A", "B", "C", "D", "E"], items):
        r[f"choice_{new_letter.lower()}"] = text
        if text == correct_text:
            r["correct_answer"] = new_letter


# ── Content hash (identical-question detection) ───────────────────────────────
def _content_hash(stimulus: str, choice_a: str = "", choice_b: str = "",
                   choice_c: str = "", choice_d: str = "",
                   choice_e: str = "", correct_answer: str = "") -> str:
    parts = [stimulus, choice_a, choice_b, choice_c,
             choice_d, choice_e, correct_answer]
    blob = "|".join(p.strip().lower() for p in parts)
    return hashlib.sha256(blob.encode()).hexdigest()[:32]


def _read_upload(file_obj) -> pd.DataFrame:
    name = getattr(file_obj, "name", "") or ""
    is_excel = name.lower().endswith((".xlsx", ".xlsm", ".xls"))

    if not is_excel and hasattr(file_obj, "tell") and hasattr(file_obj, "read"):
        pos = file_obj.tell()
        signature = file_obj.read(4)
        file_obj.seek(pos)
        is_excel = signature == b"PK\x03\x04"

    if is_excel:
        return pd.read_excel(file_obj, dtype=str).fillna("")
    return pd.read_csv(file_obj, dtype=str).fillna("")


def load_csv(file_obj) -> tuple:
    """
    Parse an uploaded file object.
    Returns (rows, errors, rows_read).
      rows      — list of valid row dicts (with content_hash added)
      errors    — list of human-readable error strings
      rows_read — total data rows seen in the file (before validation)
    """
    errors: list = []

    try:
        df = _read_upload(file_obj)
    except Exception as e:
        return [], [f"Could not parse CSV: {e}"], 0

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        return [], [f"Missing required columns: {', '.join(sorted(missing))}"], 0

    for col in ALL_COLS:
        if col not in df.columns:
            df[col] = ""

    rows_read = len(df)
    rows: list = []

    for idx, row in df.iterrows():
        line = idx + 2
        r    = {col: str(row.get(col, "")).strip() for col in ALL_COLS}

        is_open_ended = is_open_ended_question(r)

        # Validate answer fields. Multiple-choice questions need A-E choices and
        # a letter answer; open-ended questions may leave choices blank and use
        # correct_answer as an optional sample answer or rubric.
        if not is_open_ended:
            missing_choices = [
                letter.upper()
                for letter in ["a", "b", "c", "d", "e"]
                if not r[f"choice_{letter}"]
            ]
            if missing_choices:
                errors.append(
                    f"Row {line}: multiple-choice questions need choices "
                    f"{', '.join(missing_choices)}."
                )
                continue

        if not is_open_ended and r["correct_answer"].upper() not in VALID_ANSWERS:
            errors.append(
                f"Row {line}: invalid correct_answer '{r['correct_answer']}'. "
                "Must be A, B, C, D, or E for multiple-choice questions."
            )
            continue

        # Validate / default difficulty
        if r["difficulty"] == "":
            r["difficulty"] = 3
        else:
            try:
                diff = int(r["difficulty"])
                if diff < 1 or diff > 5:
                    raise ValueError
                r["difficulty"] = diff
            except (ValueError, TypeError):
                errors.append(
                    f"Row {line}: difficulty must be 1–5, "
                    f"got '{r['difficulty']}'. Defaulting to 3."
                )
                r["difficulty"] = 3

        if is_open_ended:
            r["correct_answer"] = r["correct_answer"].strip()
        else:
            r["correct_answer"] = r["correct_answer"].upper()
            _shuffle_multiple_choice_row(r)

        # Incoming IDs are treated as source labels only. The database assigns
        # the course-scoped master question_id during insert.
        r["question_id"] = ""

        # Compute content_hash for duplicate detection
        r["content_hash"] = _content_hash(
            r["stimulus"],
            r["choice_a"], r["choice_b"], r["choice_c"],
            r["choice_d"], r["choice_e"], r["correct_answer"],
        )

        rows.append(r)

    return rows, errors, rows_read


def process_upload(file_obj, course_id=None) -> dict:
    """
    Full pipeline: parse CSV → validate → insert.

    Returns a dict:
      rows_read        — total data rows in the file
      valid_rows       — rows that passed validation
      inserted         — new questions added to the database
      skipped_id       — rows skipped because generated question_id already exists
      skipped_content  — rows skipped because identical content already exists
      invalid          — rows that failed validation (wrong answer, bad difficulty…)
      errors           — list of validation error strings
    """
    rows, errors, rows_read = load_csv(file_obj)
    invalid = rows_read - len(rows) if rows_read > len(rows) else len(errors)

    if not rows:
        return {
            "rows_read": rows_read,
            "valid_rows": 0,
            "inserted": 0,
            "skipped_id": 0,
            "skipped_content": 0,
            "invalid": invalid,
            "errors": errors,
        }

    _apply_source_module_rules(rows, course_id=course_id)
    inserted, skipped_id, skipped_content = insert_questions(rows, course_id=course_id)

    return {
        "rows_read": rows_read,
        "valid_rows": len(rows),
        "inserted": inserted,
        "skipped_id": skipped_id,
        "skipped_content": skipped_content,
        "invalid": invalid,
        "errors": errors,
    }


def _template_rows() -> list[dict]:
    return [
        {
            "question_id":   "source_001_ignored",
            "section_type":  "Quantitative Methods",
            "question_type": "Multiple Choice",
            "difficulty":    1,
            "passage":       "",
            "stimulus": (
                "A bond is priced at 98. If required yields fall slightly, "
                "what should happen to the bond price?"
            ),
            "choice_a":      "It should rise.",
            "choice_b":      "It should fall.",
            "choice_c":      "It should stay exactly 98.",
            "choice_d":      "It should become par immediately.",
            "choice_e":      "It cannot be estimated from yield direction.",
            "correct_answer": "A",
            "explanation": (
                "Bond prices and required yields move in opposite directions."
            ),
            "wrong_answer_a": "",
            "wrong_answer_b": "",
            "wrong_answer_c": "",
            "wrong_answer_d": "",
            "wrong_answer_e": "",
            "source": "generated_workbook",
            "tags":   "intuition, estimation, bonds",
        },
        {
            "question_id":   "source_002_ignored",
            "section_type":  "Writing Practice",
            "question_type": "Open-Ended",
            "difficulty":    3,
            "passage":       "",
            "stimulus":      "Explain why diversification can reduce portfolio risk without eliminating all risk.",
            "choice_a":      "",
            "choice_b":      "",
            "choice_c":      "",
            "choice_d":      "",
            "choice_e":      "",
            "correct_answer": "A strong response distinguishes diversifiable risk from market risk.",
            "explanation":   "Open-ended responses are saved as written answers; this field can hold a sample answer or rubric.",
            "wrong_answer_a": "",
            "wrong_answer_b": "",
            "wrong_answer_c": "",
            "wrong_answer_d": "",
            "wrong_answer_e": "",
            "source": "generated_workbook",
            "tags":   "written response, diversification, risk",
        },
    ]


def make_template_csv() -> str:
    """Return a CSV string with headers and two example rows."""
    sample = pd.DataFrame(_template_rows(), columns=ALL_COLS)
    return sample.to_csv(index=False)


def make_template_xlsx() -> bytes:
    """Return an Excel template with validation for question_type."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Question Bank Template"

    for col_idx, header in enumerate(ALL_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="365F91")
        ws.column_dimensions[cell.column_letter].width = max(14, min(42, len(header) + 4))

    for row_idx, row in enumerate(_template_rows(), start=2):
        for col_idx, header in enumerate(ALL_COLS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    question_type_col = ALL_COLS.index("question_type") + 1
    question_type_letter = ws.cell(row=1, column=question_type_col).column_letter
    validation = DataValidation(
        type="list",
        formula1='"Multiple Choice,Open-Ended"',
        allow_blank=False,
    )
    validation.error = "Choose either Multiple Choice or Open-Ended."
    validation.errorTitle = "Invalid question type"
    validation.prompt = "Select Multiple Choice or Open-Ended."
    validation.promptTitle = "Question type"
    ws.add_data_validation(validation)
    validation.add(f"{question_type_letter}2:{question_type_letter}1000")

    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
